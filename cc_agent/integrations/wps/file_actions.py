from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

from cc_agent.domain.enums import RunStatus
from cc_agent.domain.models import StepResult, TaskStep
from cc_agent.integrations.base import BaseAdapter


class WpsAdapter(BaseAdapter):
    name = "wps"

    def __init__(self, config: dict[str, Any], exports_dir: Path) -> None:
        self._config = config
        self._exports_dir = exports_dir

    @property
    def capabilities(self) -> list[str]:
        return ["inspect_workbook", "count_duplicates_in_column", "handle_document_request"]

    def execute(self, step: TaskStep) -> StepResult:
        workbook_path = str(step.params.get("workbook_path", "")).strip()
        if step.action == "inspect_workbook" and workbook_path:
            path = Path(workbook_path)
            workbook = load_workbook(path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return StepResult(
                step_id=step.id,
                status=RunStatus.SUCCEEDED,
                message="\u5de5\u4f5c\u7c3f\u68c0\u67e5\u5b8c\u6210\u3002",
                data={"workbook_path": str(path), "sheet_names": sheet_names},
            )
        if step.action == "count_duplicates_in_column" and workbook_path:
            return self._count_duplicates(step)
        return StepResult(
            step_id=step.id,
            status=RunStatus.SUCCEEDED,
            message="WPS \u9002\u914d\u5668\u5df2\u63a5\u6536\u8be5\u8bf7\u6c42\u3002",
            data={"command": step.params.get("command", "")},
        )

    def _count_duplicates(self, step: TaskStep) -> StepResult:
        workbook_path = Path(str(step.params["workbook_path"]))
        sheet_name = str(step.params.get("sheet_name", "")).strip()
        column = str(step.params.get("column", "A")).upper()
        skip_header = bool(step.params.get("skip_header", True))

        workbook = load_workbook(workbook_path, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            resolved_sheet_name = worksheet.title
            column_index = column_index_from_string(column)
            start_row = 2 if skip_header else 1
            counts: Counter[str] = Counter()

            for row_index in range(start_row, worksheet.max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                normalized = str(value).strip()
                if normalized:
                    counts[normalized] += 1
        finally:
            workbook.close()

        duplicates = sorted(
            ((value, count) for value, count in counts.items() if count > 1),
            key=lambda item: (-item[1], item[0]),
        )

        output_path = str(step.params.get("output_path", "")).strip()
        if not output_path:
            output_path = str(self._exports_dir / f"{workbook_path.stem}_duplicates_{column}.xlsx")
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        report = Workbook()
        report_sheet = report.active
        report_sheet.title = "\u91cd\u590d\u7edf\u8ba1"
        report_sheet.append(["\u503c", "\u6b21\u6570"])
        for value, count in duplicates:
            report_sheet.append([value, count])
        report.save(output_path)
        report.close()

        return StepResult(
            step_id=step.id,
            status=RunStatus.SUCCEEDED,
            message="\u91cd\u590d\u503c\u7edf\u8ba1\u62a5\u544a\u5df2\u5bfc\u51fa\u3002",
            data={
                "workbook_path": str(workbook_path),
                "sheet_name": sheet_name or resolved_sheet_name,
                "column": column,
                "duplicate_value_count": len(duplicates),
                "output_path": output_path,
            },
            artifacts=[output_path],
        )
