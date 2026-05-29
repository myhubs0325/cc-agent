from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from cc_agent.domain.errors import AgentError
from cc_agent.onboarding.field_mapper import OnboardingFieldMapper
from cc_agent.onboarding.profile_schema import OnboardingFieldCandidate, OnboardingProfile, ParsedOnboardingSource
from cc_agent.onboarding.template_parser import OnboardingTemplateParser


class OnboardingDocumentError(AgentError):
    """Raised when an onboarding source cannot be parsed."""


class OnboardingDocumentParser:
    def __init__(
        self,
        template_parser: OnboardingTemplateParser | None = None,
        field_mapper: OnboardingFieldMapper | None = None,
    ) -> None:
        self._template_parser = template_parser or OnboardingTemplateParser()
        self._field_mapper = field_mapper or OnboardingFieldMapper()

    def parse(self, source_path: str | Path) -> ParsedOnboardingSource:
        path = Path(source_path).expanduser().resolve()
        if not path.exists():
            raise OnboardingDocumentError(f"找不到资料来源: {path}")
        if path.is_dir():
            template_root = self._resolve_template_root(path)
            parsed = self._template_parser.parse(template_root)
            return ParsedOnboardingSource(
                profile=self._enrich_profile(parsed.profile),
                warnings=parsed.warnings,
                source_kind="template_directory",
                candidates=self._profile_candidates(parsed.profile),
            )

        suffix = path.suffix.lower()
        if suffix == ".csv":
            pairs = self._parse_csv(path)
        elif suffix == ".txt":
            pairs = self._parse_text(path.read_text(encoding="utf-8", errors="ignore"))
        elif suffix == ".docx":
            pairs = self._parse_docx(path)
        elif suffix == ".xlsx":
            pairs = self._parse_xlsx(path)
        else:
            raise OnboardingDocumentError(f"暂不支持的资料格式: {suffix or '<none>'}")

        if not pairs:
            raise OnboardingDocumentError(f"未能从资料中提取键值字段: {path}")

        candidates = [OnboardingFieldCandidate(source_label=label, value=value) for label, value in pairs]
        profile = self._enrich_profile(self._field_mapper.map_candidates(candidates, str(path)))
        warnings = self._build_warnings(profile, candidates)
        return ParsedOnboardingSource(
            profile=profile,
            warnings=warnings,
            source_kind=suffix.lstrip("."),
            candidates=candidates,
        )

    def _resolve_template_root(self, path: Path) -> Path:
        try:
            self._template_parser.parse(path)
            return path
        except Exception:
            pass
        for child in sorted(path.iterdir()):
            if not child.is_dir():
                continue
            try:
                self._template_parser.parse(child)
                return child
            except Exception:
                continue
        return path

    @staticmethod
    def _profile_candidates(profile: OnboardingProfile) -> list[OnboardingFieldCandidate]:
        candidates: list[OnboardingFieldCandidate] = []
        for label, value in (
            ("公司名称", profile.company_name or profile.customer_name or ""),
            ("Dbit 登录账号", profile.dbit_username or ""),
            ("AdsPower 登录账号", profile.adspower_username or ""),
            ("平台", profile.platform or ""),
            ("安装目录", profile.install_path or ""),
        ):
            if value:
                candidates.append(OnboardingFieldCandidate(source_label=label, value=value))
        for detail in profile.clone_details[:3]:
            if detail.clone_name:
                candidates.append(OnboardingFieldCandidate(source_label="分身名称", value=detail.clone_name))
            if detail.proxy_host:
                candidates.append(OnboardingFieldCandidate(source_label="代理 IP", value=detail.proxy_host))
        return candidates

    @staticmethod
    def _enrich_profile(profile: OnboardingProfile) -> OnboardingProfile:
        if not profile.platform and profile.clone_details:
            profile.platform = profile.clone_details[0].platform or None
        if not profile.platform_username and profile.clone_details:
            profile.platform_username = profile.clone_details[0].platform_account
        if not profile.platform_password and profile.clone_details:
            profile.platform_password = profile.clone_details[0].platform_password
        if not profile.proxy_host and profile.clone_details:
            profile.proxy_host = profile.clone_details[0].proxy_host
            profile.proxy_port = profile.clone_details[0].proxy_port
            profile.proxy_username = profile.clone_details[0].proxy_username
            profile.proxy_password = profile.clone_details[0].proxy_password
        if not profile.environment_name and profile.clone_details:
            profile.environment_name = profile.clone_details[0].adspower_environment_name or profile.clone_details[0].clone_name
        if not profile.search_keywords and profile.business_scenarios:
            profile.search_keywords = profile.business_scenarios[0].keywords
        if not profile.target_audience and profile.business_scenarios:
            profile.target_audience = profile.business_scenarios[0].target_audience
        if not profile.business_background and profile.business_scenarios:
            profile.business_background = profile.business_scenarios[0].business_intro
        if not profile.business_selling_points and profile.business_scenarios:
            profile.business_selling_points = "; ".join(profile.business_scenarios[0].selling_points)
        return profile

    @staticmethod
    def _build_warnings(
        profile: OnboardingProfile,
        candidates: list[OnboardingFieldCandidate],
    ) -> list[str]:
        warnings: list[str] = []
        if not candidates:
            warnings.append("资料中没有可识别的字段。")
        required_fields = {
            "Dbit 登录账号": profile.dbit_username,
            "Dbit 登录密码": profile.dbit_password,
            "AdsPower 登录账号": profile.adspower_username,
            "AdsPower 登录密码": profile.adspower_password,
        }
        for label, value in required_fields.items():
            if not value:
                warnings.append(f"缺少关键字段: {label}")
        if not profile.proxy_host or not profile.proxy_port:
            warnings.append("代理信息未填写完整，后续流程会停在代理等待阶段。")
        return warnings

    @staticmethod
    def _parse_csv(path: Path) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                values = [str(item).strip() for item in row if str(item).strip()]
                if len(values) >= 2:
                    pairs.append((values[0], values[1]))
                elif values:
                    pairs.extend(OnboardingDocumentParser._parse_text(values[0]))
        return pairs

    @staticmethod
    def _parse_text(text: str) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            match = re.match(r"^\s*([^:：,，]+)\s*[:：,，]\s*(.+?)\s*$", line)
            if match:
                pairs.append((match.group(1).strip(), match.group(2).strip()))
        return pairs

    @staticmethod
    def _parse_docx(path: Path) -> list[tuple[str, str]]:
        paragraphs: list[str] = []
        with zipfile.ZipFile(path) as archive:
            try:
                xml_bytes = archive.read("word/document.xml")
            except KeyError as exc:
                raise OnboardingDocumentError(f"docx 缺少正文 XML: {path}") from exc
        root = ElementTree.fromstring(xml_bytes)
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for paragraph in root.findall(".//w:p", namespace):
            texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
            content = "".join(texts).strip()
            if content:
                paragraphs.append(content)
        return OnboardingDocumentParser._parse_text("\n".join(paragraphs))

    @staticmethod
    def _parse_xlsx(path: Path) -> list[tuple[str, str]]:
        workbook = load_workbook(filename=path, data_only=True, read_only=True)
        pairs: list[tuple[str, str]] = []
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if len(values) >= 2:
                        pairs.append((values[0], values[1]))
                    elif len(values) == 1:
                        pairs.extend(OnboardingDocumentParser._parse_text(values[0]))
        finally:
            workbook.close()
        return pairs
