from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from cc_agent.domain.errors import AgentError
from cc_agent.onboarding.profile_schema import (
    BusinessScenario,
    ChecklistItem,
    CloneAccountDetail,
    CloneResourcePlan,
    OnboardingProfile,
    ParsedOnboardingTemplate,
)


class OnboardingTemplateError(AgentError):
    """Raised when the onboarding template cannot be parsed."""


_REQUIRED_FILE_STEMS = {
    "base_info": "01基础资料",
    "clone_plan": "02分身数量规划",
    "clone_details": "03分身明细",
    "business_scenarios": "04关键词与话术",
    "checklist": "05交付前检查",
}

_SUPPORTED_TEMPLATE_SUFFIXES = (".csv", ".xlsx", ".xlsm")

_BASE_INFO_FIELD_MAP = {
    "公司/品牌名称": "company_name",
    "联系人姓名": "contact_name",
    "联系人手机/微信": "contact_handle",
    "你主要卖什么": "business_summary",
    "这次主要做哪个国家/地区": "target_market",
    "TikTok 计划做几个分身": "tiktok_clone_count",
    "Facebook 计划做几个分身": "facebook_clone_count",
    "预计要准备几个 AdsPower 环境": "planned_adspower_environment_count",
    "Dbit 系统登录账号": "dbit_username",
    "Dbit 系统登录密码": "dbit_password",
    "Dbit 系统授权码": "dbit_license_code",
    "AdsPower 登录账号": "adspower_username",
    "AdsPower 登录密码": "adspower_password",
    "AdsPower 是否已开通专业版及以上": "adspower_plan_ready",
    "AI 模型 API Key": "ai_model_api_key",
    "代理类型": "proxy_mode",
    "代理用户ID": "proxy_user_id",
    "代理密钥": "proxy_sync_key",
    "软件准备安装到哪里": "install_path",
    "自动发帖素材所在文件夹": "asset_folder",
    "补充备注": "notes",
}


class OnboardingTemplateParser:
    def parse(self, template_dir: str | Path) -> ParsedOnboardingTemplate:
        root = Path(template_dir).expanduser().resolve()
        if not root.exists() or not root.is_dir():
            raise OnboardingTemplateError(f"找不到资料模板目录: {root}")

        file_map = self._resolve_required_files(root)
        base_rows = self._read_rows(file_map["base_info"])
        plan_rows = self._read_rows(file_map["clone_plan"])
        detail_rows = self._read_rows(file_map["clone_details"])
        scenario_rows = self._read_rows(file_map["business_scenarios"])
        checklist_rows = self._read_rows(file_map["checklist"])

        profile = self._build_profile(root, base_rows, plan_rows, detail_rows, scenario_rows, checklist_rows)
        warnings = self._collect_warnings(profile)
        return ParsedOnboardingTemplate(profile=profile, warnings=warnings)

    def _resolve_required_files(self, root: Path) -> dict[str, Path]:
        resolved: dict[str, Path] = {}
        missing: list[str] = []
        for key, stem in _REQUIRED_FILE_STEMS.items():
            path = self._find_template_file(root, stem)
            if path is None:
                missing.append(f"{stem}.csv/.xlsx")
                continue
            resolved[key] = path
        if missing:
            raise OnboardingTemplateError(f"资料模板目录缺少必需文件: {'、'.join(missing)}")
        return resolved

    def _find_template_file(self, root: Path, stem: str) -> Path | None:
        for suffix in _SUPPORTED_TEMPLATE_SUFFIXES:
            candidate = root / f"{stem}{suffix}"
            if candidate.exists():
                return candidate
        return None

    def _read_rows(self, path: Path) -> list[dict[str, str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._read_csv_rows(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx_rows(path)
        raise OnboardingTemplateError(f"不支持的模板文件格式: {path.name}")

    @staticmethod
    def _read_csv_rows(path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return [{(key or "").strip(): (value or "").strip() for key, value in row.items()} for row in reader]

    @staticmethod
    def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(filename=path, data_only=True, read_only=True)
        rows: list[dict[str, str]] = []
        try:
            for worksheet in workbook.worksheets:
                header: list[str] | None = None
                for raw_row in worksheet.iter_rows(values_only=True):
                    values = [("" if cell is None else str(cell).strip()) for cell in raw_row]
                    if not any(values):
                        continue
                    if header is None:
                        header = values
                        continue
                    row = {
                        (header[index] if index < len(header) else "").strip(): value
                        for index, value in enumerate(values)
                        if (header[index] if index < len(header) else "").strip()
                    }
                    if any(value for value in row.values()):
                        rows.append(row)
        finally:
            workbook.close()
        return rows

    def _build_profile(
        self,
        root: Path,
        base_rows: list[dict[str, str]],
        plan_rows: list[dict[str, str]],
        detail_rows: list[dict[str, str]],
        scenario_rows: list[dict[str, str]],
        checklist_rows: list[dict[str, str]],
    ) -> OnboardingProfile:
        base_data = self._parse_base_info(base_rows)
        clone_plans = self._parse_clone_plans(plan_rows)
        clone_details = self._parse_clone_details(detail_rows)
        business_scenarios = self._parse_business_scenarios(scenario_rows)
        checklist = self._parse_checklist(checklist_rows)

        requested_platforms = _split_tokens(
            base_data.pop("requested_platforms_raw", ""),
            separators=r"[,，/\+\n、]",
        )
        if not requested_platforms:
            requested_platforms = sorted({item.platform for item in clone_details if item.platform})

        return OnboardingProfile(
            template_dir=str(root),
            requested_platforms=requested_platforms,
            clone_plans=clone_plans,
            clone_details=clone_details,
            business_scenarios=business_scenarios,
            checklist=checklist,
            **base_data,
        )

    def _parse_base_info(self, rows: list[dict[str, str]]) -> dict[str, object]:
        data: dict[str, object] = {"requested_platforms_raw": ""}
        for row in rows:
            item = row.get("项目", "")
            value = row.get("请填写", "")
            if not item:
                continue
            if item == "这次要做哪些平台":
                data["requested_platforms_raw"] = value
                continue
            field_name = _BASE_INFO_FIELD_MAP.get(item)
            if not field_name:
                continue
            if field_name in {"tiktok_clone_count", "facebook_clone_count", "planned_adspower_environment_count"}:
                data[field_name] = _to_int(value) or 0
                continue
            if field_name == "adspower_plan_ready":
                data[field_name] = _to_bool(value)
                continue
            data[field_name] = value or None
        return data

    @staticmethod
    def _parse_clone_plans(rows: list[dict[str, str]]) -> list[CloneResourcePlan]:
        plans: list[CloneResourcePlan] = []
        for row in rows:
            platform = row.get("平台", "")
            if not platform or platform == "合计":
                continue
            plans.append(
                CloneResourcePlan(
                    platform=platform,
                    clone_count=_to_int(row.get("计划做几个分身", "")),
                    account_count=_to_int(row.get("建议准备几个平台账号", "")),
                    proxy_count=_to_int(row.get("建议准备几条代理 IP", "")),
                    environment_count=_to_int(row.get("建议准备几个 AdsPower 环境", "")),
                    note=row.get("说明", "") or None,
                )
            )
        return plans

    @staticmethod
    def _parse_clone_details(rows: list[dict[str, str]]) -> list[CloneAccountDetail]:
        details: list[CloneAccountDetail] = []
        for row in rows:
            sequence = row.get("序号", "")
            platform = row.get("平台", "")
            clone_name = row.get("分身名称", "")
            if not any(row.values()):
                continue
            if sequence == "示例":
                continue
            if not platform and not clone_name:
                continue
            details.append(
                CloneAccountDetail(
                    sequence=sequence or "",
                    platform=platform or "",
                    clone_name=clone_name or None,
                    adspower_environment_name=row.get("AdsPower 环境名称", "") or None,
                    platform_account=row.get("平台账号/邮箱/手机号", "") or None,
                    platform_password=row.get("平台密码", "") or None,
                    account_region=row.get("账号国家/地区", "") or None,
                    proxy_host=row.get("代理 IP", "") or None,
                    proxy_port=row.get("代理端口", "") or None,
                    proxy_username=row.get("代理用户名", "") or None,
                    proxy_password=row.get("代理密码/密钥", "") or None,
                    login_verified=_to_bool(row.get("账号是否已手动登录测试", "")),
                    profile_completed=_to_bool(row.get("头像昵称资料是否已完善", "")),
                    note=row.get("备注", "") or None,
                )
            )
        return details

    @staticmethod
    def _parse_business_scenarios(rows: list[dict[str, str]]) -> list[BusinessScenario]:
        scenarios: list[BusinessScenario] = []
        for row in rows:
            scenario_name = row.get("业务场景名称", "")
            if not scenario_name:
                continue
            if scenario_name.startswith("示例："):
                continue
            scenarios.append(
                BusinessScenario(
                    scenario_name=scenario_name,
                    platform=row.get("适用平台", "") or None,
                    reuse_clone_count=_to_int(row.get("准备给几个分身复用", "")),
                    keywords=_split_tokens(row.get("搜索关键词（英文优先，多个词用逗号隔开）", "")),
                    target_audience=row.get("目标客户画像", "") or None,
                    excluded_audience=row.get("不想要的客户/过滤条件", "") or None,
                    business_intro=row.get("用一句话介绍你是做什么的", "") or None,
                    selling_points=_split_tokens(row.get("你的 3-5 条核心卖点", ""), separators=r"[,\n;；、]"),
                    comment_reply_template=row.get("评论区 1 对 1 回复模板（可选）", "") or None,
                    auto_post_copy=row.get("自动发帖文案（可选）", "") or None,
                    asset_folder=row.get("图片/视频素材文件夹（可选）", "") or None,
                )
            )
        return scenarios

    @staticmethod
    def _parse_checklist(rows: list[dict[str, str]]) -> list[ChecklistItem]:
        items: list[ChecklistItem] = []
        for row in rows:
            item = row.get("检查项", "")
            if not item:
                continue
            items.append(
                ChecklistItem(
                    item=item,
                    prepared=_to_bool(row.get("是否已准备", "")),
                    note=row.get("不会操作时就按这个做", "") or None,
                )
            )
        return items

    def _collect_warnings(self, profile: OnboardingProfile) -> list[str]:
        warnings: list[str] = []

        required_fields = {
            "公司/品牌名称": profile.company_name,
            "联系人姓名": profile.contact_name,
            "联系人手机/微信": profile.contact_handle,
            "Dbit 系统登录账号": profile.dbit_username,
            "Dbit 系统登录密码": profile.dbit_password,
            "软件安装目录": profile.install_path,
        }
        for label, value in required_fields.items():
            if not value:
                warnings.append(f"基础资料缺少必填项: {label}")

        if not profile.adspower_username or not profile.adspower_password:
            warnings.append("AdsPower 登录信息未填写，流程会进入 AdsPower 注册/登录引导步骤。")

        expected_clone_count = max(profile.tiktok_clone_count, 0) + max(profile.facebook_clone_count, 0)
        if expected_clone_count <= 0:
            warnings.append("基础资料中的分身数量还没有填写，后续无法自动规划账号和环境。")

        if profile.planned_adspower_environment_count and expected_clone_count:
            if profile.planned_adspower_environment_count != expected_clone_count:
                warnings.append(
                    "基础资料中的 AdsPower 环境数量与 TikTok/Facebook 分身总数不一致，"
                    "建议按 1 个分身 = 1 个 AdsPower 环境重新核对。"
                )

        if profile.clone_details and expected_clone_count and len(profile.clone_details) != expected_clone_count:
            warnings.append(
                f"分身明细当前填写了 {len(profile.clone_details)} 行，但基础资料合计分身数为 {expected_clone_count}。"
            )

        if not profile.clone_details:
            warnings.append("分身明细表还没有填写任何账号与代理信息。")

        static_proxy_mode = _is_static_proxy_mode(profile.proxy_mode)
        for detail in profile.clone_details:
            missing_fields: list[str] = []
            if not detail.platform:
                missing_fields.append("平台")
            if not detail.clone_name:
                missing_fields.append("分身名称")
            if not detail.adspower_environment_name:
                missing_fields.append("AdsPower 环境名称")
            if not detail.platform_account:
                missing_fields.append("平台账号")
            if not detail.platform_password:
                missing_fields.append("平台密码")
            if not static_proxy_mode:
                if not detail.proxy_host:
                    missing_fields.append("代理 IP")
                if not detail.proxy_port:
                    missing_fields.append("代理端口")
            if missing_fields:
                title = detail.clone_name or detail.sequence or "未命名分身"
                warnings.append(f"分身明细 [{title}] 缺少字段: {'、'.join(missing_fields)}")

        if static_proxy_mode:
            if not profile.proxy_user_id:
                warnings.append("代理类型为静态，但基础资料缺少代理用户ID。")
            if not profile.proxy_sync_key:
                warnings.append("代理类型为静态，但基础资料缺少代理密钥。")
        else:
            any_proxy = any(detail.proxy_host and detail.proxy_port for detail in profile.clone_details)
            if not any_proxy and not (profile.proxy_host and profile.proxy_port):
                warnings.append("当前代理类型需要直接代理参数，但还没有填写代理 IP 和端口。")

        if not profile.business_scenarios:
            warnings.append("关键词与话术表还没有填写任何业务场景。")
        else:
            for scenario in profile.business_scenarios:
                if not scenario.keywords:
                    warnings.append(f"业务场景 [{scenario.scenario_name}] 还没有填写搜索关键词。")
                if not scenario.business_intro:
                    warnings.append(f"业务场景 [{scenario.scenario_name}] 还没有填写一句话业务介绍。")
                if not scenario.selling_points:
                    warnings.append(f"业务场景 [{scenario.scenario_name}] 还没有填写核心卖点。")

        if profile.adspower_plan_ready is False:
            warnings.append("AdsPower 套餐当前标记为未开通，后续自动化配置会停在付款等待步骤。")

        return warnings


def _to_int(value: str) -> int | None:
    normalized = (value or "").strip()
    if not normalized:
        return None
    try:
        return int(float(normalized))
    except ValueError:
        return None


def _to_bool(value: str) -> bool | None:
    normalized = (value or "").strip().lower()
    if not normalized:
        return None
    true_tokens = {"是", "已", "已准备", "已开通", "true", "yes", "y", "1"}
    false_tokens = {"否", "未", "未准备", "未开通", "false", "no", "n", "0"}
    if normalized in true_tokens:
        return True
    if normalized in false_tokens:
        return False
    return None


def _split_tokens(value: str, separators: str = r"[,，\n]") -> list[str]:
    normalized = (value or "").strip()
    if not normalized:
        return []
    return [item.strip() for item in re.split(separators, normalized) if item.strip()]


def _is_static_proxy_mode(value: str | None) -> bool:
    normalized = str(value or "").strip().casefold()
    return normalized in {"static", "static_sync", "静态", "静态代理"}
